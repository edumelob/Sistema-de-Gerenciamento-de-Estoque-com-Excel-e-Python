import openpyxl

CAMINHO_ARQUIVO = "produtos.xlsx"
ABA = "Estoque"

def carregar_planilha():
    wb = openpyxl.load_workbook(CAMINHO_ARQUIVO)
    return wb, wb[ABA]

def salvar_planilha(wb):
    wb.save(CAMINHO_ARQUIVO)

def listar_produtos(aba):
    print("\n📦 Lista de Produtos:")
    for linha in aba.iter_rows(min_row=2, values_only=True):
        print(f"ID: {linha[0]}, Produto: {linha[1]}, Qtd: {linha[3]}, Preço: R${linha[4]}")

def adicionar_produto(aba, nome, categoria, quantidade, preco):
    nova_linha = [aba.max_row, nome, categoria, quantidade, preco, quantidade * preco]
    aba.append(nova_linha)
    print(f"✅ Produto '{nome}' adicionado!")

def atualizar_quantidade(aba, id_produto, nova_quantidade):
    for linha in aba.iter_rows(min_row=2):
        if linha[0].value == id_produto:
            linha[3].value = nova_quantidade
            linha[5].value = nova_quantidade * linha[4].value
            print(f"✏️ Quantidade do produto ID {id_produto} atualizada para {nova_quantidade}")
            return
    print("❌ Produto não encontrado.")

def menu():
    wb, aba = carregar_planilha()
    while True:
        print("\n[1] Listar produtos\n[2] Adicionar produto\n[3] Atualizar quantidade\n[4] Sair")
        opcao = input("Escolha: ")

        if opcao == "1":
            listar_produtos(aba)

        elif opcao == "2":
            nome = input("Nome do produto: ")
            categoria = input("Categoria: ")
            quantidade = int(input("Quantidade: "))
            preco = float(input("Preço unitário: "))
            adicionar_produto(aba, nome, categoria, quantidade, preco)

        elif opcao == "3":
            id_produto = int(input("ID do produto: "))
            nova_qtd = int(input("Nova quantidade: "))
            atualizar_quantidade(aba, id_produto, nova_qtd)

        elif opcao == "4":
            salvar_planilha(wb)
            print("📁 Alterações salvas. Saindo...")
            break
        else:
            print("⚠️ Opção inválida!")

if __name__ == "__main__":
    menu()
